
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# Page configuration
st.set_page_config(
    page_title="Workpapers QC Dashboard",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Workpapers Quality Control Dashboard")
st.markdown("Compare measures across cost, performance, and energy metrics")

# Load data function (replace with actual data loading)
@st.cache_data
def load_data():
    # TODO: Replace this with actual data loading from workpapers
    # For now, return empty dataframe structure
    import openpyxl
    
    workbook_path = './001_input/workpapers_testing.xlsx'
    wb = openpyxl.load_workbook(workbook_path)
    
    all_sheets_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_dict = {'sheet_name': sheet_name}
        
        for row in range(2, min(102, ws.max_row + 1)):
            field_cell = ws.cell(row=row, column=1)
            value_cell = ws.cell(row=row, column=2)
            
            if field_cell.value:
                sheet_dict[field_cell.value] = value_cell.value
        
        all_sheets_data.append(sheet_dict)
    
    df = pd.DataFrame(all_sheets_data)
    
    # Convert to snake_case
    def to_snake_case(text):
        if text is None or pd.isna(text):
            return text
        text = str(text)
        text = text.lower().replace(' ', '_').replace('$', 'usd').replace('/', 'per').replace('&', 'and').replace('.', '')
        while '__' in text:
            text = text.replace('__', '_')
        return text
    
    df.columns = [to_snake_case(col) for col in df.columns]
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].apply(to_snake_case)
    
    return df

df = load_data()

# Define metric options for comparison
metric_options = {
    'Equipment Cost': 'equipment_cost',
    'Labor Cost Installation': 'labor_cost_installation',
    'Effective Useful Life (years)': 'effective_useful_life_yrs',
    'Water Usage (gal)': 'water_usage_gal',
    'O&M Costs': 'o_and_m_costs',
    'RET Retirement Rate': 'ret_retirement_rate'
}

# Define filter columns
filter_columns = [
    'primary_fuel', 'efficiency_description', 'efficiency_level', 'sector',
    'unit_of_characterization', 'competition_group', 'heating_or_cooling',
    'subgroup', 'base_building_type', 'interaction_group', 'climate_zone',
    'ret_add_on_applicable', 'ret_er_applicable', 'nc_applicable',
    'rob_applicable', 'reno_applicable', 'demo_applicable'
]

# Sidebar filters
st.sidebar.header("🔍 Filters")

# Create filters dynamically
filters = {}
for filter_col in filter_columns:
    if filter_col in df.columns:
        unique_values = df[filter_col].dropna().unique()
        if len(unique_values) > 0:
            filters[filter_col] = st.sidebar.multiselect(
                filter_col.replace('_', ' ').title(),
                options=sorted(unique_values),
                default=None
            )

# Apply filters
df_filtered = df.copy()
for filter_col, selected_values in filters.items():
    if selected_values:
        df_filtered = df_filtered[df_filtered[filter_col].isin(selected_values)]

st.sidebar.markdown(f"**Measures shown:** {len(df_filtered)} / {len(df)}")

# Main content
st.header("📈 Cost & Performance Metrics")

# Metric selector
selected_metric_name = st.selectbox(
    "Select metric to compare:",
    options=list(metric_options.keys())
)
selected_metric = metric_options[selected_metric_name]

# Check if metric exists in dataframe
if selected_metric in df_filtered.columns:
    # Convert to numeric
    df_filtered[selected_metric] = pd.to_numeric(df_filtered[selected_metric], errors='coerce')
    
    # Remove rows with null values for the selected metric
    df_plot = df_filtered[df_filtered[selected_metric].notna()].copy()
    
    if len(df_plot) > 0:
        # Sort by metric value
        df_plot = df_plot.sort_values(selected_metric, ascending=False)
        
        # Create bar chart
        fig = px.bar(
            df_plot,
            x='condition_name' if 'condition_name' in df_plot.columns else 'sheet_name',
            y=selected_metric,
            title=f"{selected_metric_name} by Measure",
            labels={
                selected_metric: selected_metric_name,
                'condition_name': 'Measure',
                'sheet_name': 'Measure'
            },
            color=selected_metric,
            color_continuous_scale='Viridis'
        )
        
        fig.update_layout(
            xaxis_tickangle=-45,
            height=600,
            showlegend=False
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Summary statistics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Mean", f"{df_plot[selected_metric].mean():.2f}")
        with col2:
            st.metric("Median", f"{df_plot[selected_metric].median():.2f}")
        with col3:
            st.metric("Min", f"{df_plot[selected_metric].min():.2f}")
        with col4:
            st.metric("Max", f"{df_plot[selected_metric].max():.2f}")
    else:
        st.warning(f"No data available for {selected_metric_name} with current filters")
else:
    st.error(f"Column '{selected_metric}' not found in dataframe")

# Energy/Fuel Analysis Section
st.header("⚡ Energy & Fuel Analysis")

# Get all fuel columns
fuel_columns = [col for col in df_filtered.columns if col.startswith('fuel_') and '_type' in col]
fuel_numbers = sorted(list(set([col.split('_')[1] for col in fuel_columns])))

if fuel_numbers:
    # Create dataframe for fuel analysis
    fuel_data = []
    
    for idx, row in df_filtered.iterrows():
        measure_name = row.get('condition_name', row.get('sheet_name', 'Unknown'))
        
        for fuel_num in fuel_numbers:
            fuel_type = row.get(f'fuel_{fuel_num}_type')
            fuel_consumption = row.get(f'fuel_{fuel_num}_consumption')
            fuel_units = row.get(f'fuel_{fuel_num}_units')
            fuel_end_use = row.get(f'fuel_{fuel_num}_end_use')
            
            # Only include if fuel type is not null
            if pd.notna(fuel_type) and fuel_type != '':
                fuel_data.append({
                    'measure': measure_name,
                    'fuel_number': fuel_num,
                    'fuel_type': fuel_type,
                    'consumption': pd.to_numeric(fuel_consumption, errors='coerce'),
                    'units': fuel_units,
                    'end_use': fuel_end_use
                })
    
    if fuel_data:
        df_fuel = pd.DataFrame(fuel_data)
        
        # Fuel type selector
        fuel_type_filter = st.multiselect(
            "Filter by fuel type:",
            options=sorted(df_fuel['fuel_type'].unique()),
            default=None
        )
        
        if fuel_type_filter:
            df_fuel = df_fuel[df_fuel['fuel_type'].isin(fuel_type_filter)]
        
        # Remove null consumption values
        df_fuel = df_fuel[df_fuel['consumption'].notna()]
        
        if len(df_fuel) > 0:
            # Create grouped bar chart by fuel type
            fig_fuel = px.bar(
                df_fuel,
                x='measure',
                y='consumption',
                color='fuel_type',
                title="Fuel Consumption by Measure and Fuel Type",
                labels={'consumption': 'Consumption', 'measure': 'Measure'},
                barmode='group',
                hover_data=['units', 'end_use']
            )
            
            fig_fuel.update_layout(
                xaxis_tickangle=-45,
                height=600
            )
            
            st.plotly_chart(fig_fuel, use_container_width=True)
            
            # Fuel summary table
            st.subheader("Fuel Summary")
            fuel_summary = df_fuel.groupby(['fuel_type', 'units']).agg({
                'consumption': ['count', 'sum', 'mean', 'min', 'max']
            }).round(2)
            st.dataframe(fuel_summary, use_container_width=True)
        else:
            st.info("No fuel consumption data available with current filters")
    else:
        st.info("No fuel data found in the workpapers")
else:
    st.info("No fuel columns found in the workpapers")

# Data table view
with st.expander("📋 View Filtered Data Table"):
    st.dataframe(df_filtered, use_container_width=True)
